import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime

now = datetime.now()
dt_string = now.strftime("%d-%m-%Y_%H-%M-%S")

ecomm = ['Shopify Fulfillment Network','C3 Presents-ACL Festival', 'C3 Presents-Lollapalooza', 'C3 Presents-Voodoo Festival', 'C3 Presents-Sea.Hear.Now.Festival', 'Chopfit', 'Elemental Knives', 'Hi How Are You', 'Knocki', 'Mohawk', 'Neuron Garage', 'Primal 7', 'Ranch Rider', 'Ripple Reads', 'Save Muny',"Serota's Underarm Balm",'Spellbound','SXSW','Vochill','Ubuntu', 'Waterloo Sparkling Water']
WHSL_packers = ['Logan Zimmerhanzel', 'Dal Rhoads', 'Madison  Zapata', 'Taylor Hurtado', 'Jarid Craig']

wb = load_workbook('Shipments_without_append.xlsx', data_only=True)
ws = wb['SH_Shipments']
# colQ = ws['Q']

new = Workbook()
ws1 = new.create_sheet(title="ecomm-WHSL")
row_counter = 1
col_counter = 1

# copying original workbook
for row in ws.rows:
    col_counter = 1
    for cell in row:
        ws1.cell(row=row_counter, column=col_counter).value = cell.value
        col_counter += 1
    row_counter += 1

ws1.cell(row=1,column=18).value = 'Ecomm/WHSL'
ws1.cell(row=1,column=19).value = 'Merchant Tag'
ws1.cell(row=1,column=20).value = 'Countunique Orders'
ws1.cell(row=1,column=21).value = 'Concat'
ws1.cell(row=1,column=22).value = 'Kelly Wynne Item'
ws1.cell(row=1,column=23).value = 'Items 2-5'
ws1.cell(row=1,column=24).value = 'Items 6+'
del new['Sheet']
# new.save(filename=f'testShip_{dt_string}.xlsx')
# ws1.cell(row=1, column=18).value = 'ec-wh'
# ws1.cell(row=1, column=19).value = 'block'
# colQ = ws1['Q']

# Q is 17, R is 18

row_counter = 2
col_counter = 1
maxrow = ws1.max_row

# # populating col R with ecomm or WHSL
for num in range(2,maxrow+1):
    if str(ws1.cell(row=num, column=17).value) in ecomm:
        ws1.cell(row=num, column=18).value = 'ecomm'
    elif str(ws1.cell(row=num, column=17).value) not in ecomm:
        if ws1.cell(row=num, column=6).value == 'Manual Order' and ws1.cell(row=num, column=3).value >= 10:
            ws1.cell(row=num, column=18).value = 'WHSL'
        elif ws1.cell(row=num, column=17).value == 'Kammok' and (str(ws1.cell(row=num, column=7).value)[0:2] == 'KW' or str(ws1.cell(row=num, column=7).value)[0:6] == 'KMK-HQ'):
            ws1.cell(row=num, column=18).value = 'WHSL'
        elif ws1.cell(row=num, column=17).value == 'Howler Brothers' and str(ws1.cell(row=num, column=7).value)[0:2] == 'IF':
            ws1.cell(row=num, column=18).value = 'WHSL'
        elif ws1.cell(row=num, column=17).value == 'Airhouse' and ws1.cell(row=num, column=3).value != None and ws1.cell(row=num, column=3).value > 10:
            ws1.cell(row=num, column=18).value = 'WHSL'
        elif ws1.cell(row=num, column=1).value in WHSL_packers:
            ws1.cell(row=num, column=18).value = 'WHSL'
        else:
            ws1.cell(row=num, column=18).value = 'ecomm'
    else:
        ws1.cell(row=num, column=18).value = 'ecomm'
# # print(row_counter)



# col U (21) is Q (17) + G (7)
for num in range(2,maxrow+1):
    ws1.cell(row=num, column=21).value = str(str(ws1.cell(row=num, column=17).value) + str(ws1.cell(row=num, column=7).value))

# kelly wynne
for num in range(2,maxrow+1):
    if ws1.cell(row=num, column=17).value =='Kelly Wynne':
        ws1.cell(row=num, column=22).value = 1
    else:
        ws1.cell(row=num, column=22).value = 0

# item 6+
for num in range(2,maxrow+1):
    if ws1.cell(row=num, column=3).value is None:
        ws1.cell(row=num, column=24).value = 0
    elif ws1.cell(row=num, column=3).value is not None and int(ws1.cell(row=num, column=3).value) - 5 > 0:
        ws1.cell(row=num, column=24).value = int(ws1.cell(row=num, column=3).value) - 5
    else:
        ws1.cell(row=num, column=24).value = 0

# item 2-5
for num in range(2,maxrow+1):
    if ws1.cell(row=num, column=3).value is None:
        ws1.cell(row=num, column=3).value = 0
    ws1.cell(row=num, column=23).value = int(ws1.cell(row=num, column=3).value) - int(ws1.cell(row=num, column=22).value) - int(ws1.cell(row=num, column=24).value)

new.save(filename=f'Shipments_{dt_string}.xlsx')