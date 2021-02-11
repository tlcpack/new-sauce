import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd

wb = load_workbook('Client Detail_2021_2H_concat_test.xlsx', data_only=True, read_only=True)
row_counter = 2
bold = Font(bold=True)
total_clients = ['C3 Presents-ACL Festival', 'C3 Presents-Lollapalooza', 'C3 Presents-Voodoo Festival', 'C3 Presents-Sea.Hear.Now.Festival']
# total_clients = ['Airhouse', 'Alvies', 'ArtStartArt', 'Atomic Athlete', 'Austin City Limits', 'Austin Eastciders', 'Austin Emergency Supply Foundation', 'Bumper Active', 'C3 Presents', 'ChopFit', 'Classic Dad Move', 'Department of Brewology', 'Duck Camp', 'Elemental Knives', 'Evil Nails', 'Hands To Hearts', 'HELM', 'Hi How Are You', 'Howler Brothers', 'I Build America', 'Kammok', 'Kelly Wynne', 'Knocki', 'Membersy', 'Mohawk Austin', 'Moon Fabrications', 'Neuron Garage', 'Nomad Grills', 'One Blade', 'Pepper', 'Primal 7', 'Ranch Rider', 'Rowdy Gentleman', 'Rowing Blazers', 'Save Muny', "Serota's Underarm Balm", 'Shopify Fulfillment Network', None, 'Siete', 'Spellbound Sleep', 'SXSW', 'Texas Humor', 'Texas Monthly', 'The Chivery', "Tito's Vodka", 'Trek Light Gear', 'Ubuntu', 'Vochill', 'Warstic', 'Waterloo Sparkling Water', 'William Murray Golf']
all_sheets = wb.sheetnames
# all_sheets = ['FedEx', 'USPS', 'DHL E-Commerce', 'UPS', 'APC', 'SH Product Locations', 'Pallet Counts', 'ShipHero Shipments', 'Heatonist Subscriptions', 'The Chivery', 'Returns', 'WMG Returns', 'Howler Returns', 'Purchase Orders', 'Labor']
print("loaded")

### which worksheet has client names
ws1 = wb['Weekly Pallet Counts']

### populating client list
for rowNum in range(2, 127):
    client_name = ws1.cell(row=rowNum, column=1).value
    if client_name not in total_clients:
        total_clients.append(client_name)

# for row in ws1.rows:
#     # for col in row:
#     #     print(col.value)
#     client_name = ws1.cell(row=row, column=1).value
#     if client_name not in total_clients:
#         total_clients.append(client_name)

all_clients = list(filter(None, total_clients))
# print(all_clients)

### making new WB
def new_workbook():
    created_wb = Workbook()
    for sheet in all_sheets:
        created_wb.create_sheet(title=sheet)
        for i in range(1, 2):
            for j in range(1, wb[str(sheet)].max_column + 1):
                created_wb[str(sheet)].cell(row=i, column=j).value = wb[str(sheet)].cell(row=i, column=j).value
                created_wb[str(sheet)].cell(row=i, column=j).font = bold
    del created_wb['Sheet']
    print('workbook created')
    return created_wb

# ws_max_rows = ws1.max_row
# ws_max_columns = ws1.max_column
# wb[str(sheet)].max_row + 1
# =IF(W2="C3 Presents",CONCAT(W2,"-",Z2),W2)

testing_clients=['Howler Brothers', 'William Murray Golf', 'Airhouse', 'HELM']

for client in testing_clients:
    client_wb = new_workbook()
    for sheet in all_sheets:
        if str(sheet) == 'FedEx': # W 23
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 23).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
            print('fedex')
        if str(sheet) == 'USPS': # 52 AZ
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 52).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
            print('usps')
        if str(sheet) == 'DHL E-Com': # 48 AV
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 48).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'APC': # 14
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 14).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'UPS': # 17 Q
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 17).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Shippo': # 21
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 21).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Dropoff': # 44
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 44).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'SH_Product_Locations': # 13 M
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 13).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Weekly Pallet Counts': # 4
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 4).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Shipped Items_Chivery': # 21
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 21).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Shipments_Heatonist': # 21
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 21).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Pepper Carton Count': # 2
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 2).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Labor': # 16
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 16).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'SH_Purchase_Orders': # 10
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 10).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'SH Returns': # 16 P
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 16).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Howler Returns': # 10
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 10).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'WMG Loop': # 7
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 7).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'Direct Procurement': # 8 H
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 8).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1
        if str(sheet) == 'SH_Shipments': # 20 T
            actual_row = 1
            row_counter = 2
            for row in wb[str(sheet)].rows:
                actual_col = 1
                if client == wb[str(sheet)].cell(row= actual_row, column= 20).value:
                    for cell in row:
                        client_wb[str(sheet)].cell(row = row_counter, column = actual_col).value = cell.value
                        actual_col += 1
                    row_counter += 1
                actual_row += 1     
    for sheet in all_sheets:
        if client_wb[str(sheet)]['a2'].value == None and client_wb[str(sheet)]['b2'].value == None:
            del client_wb[str(sheet)]
    if len(client_wb.sheetnames) == 0:
        client_wb.create_sheet('NONE')
    client_wb.save(filename=f'Customer_{client}_test_withmax.xlsx')